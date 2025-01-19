#!/usr/bin/env python
# coding: utf-8

# In[1]:


from openpyxl import load_workbook
import re


# In[2]:


class Convert:
    def egress (self, namespace, name_policy,  number, proto, app, endpoin, port):
        if (endpoin.endswith("/32") == True):
            # Значение является CIDR
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toCIDRSet:\n"  + 8*"\xa0" + f"- cidr: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
        elif (endpoin.find("gazprom-neft.local") == True):
            # Значение является FQDN
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toFQDNs:\n"  + 8*"\xa0" + f"- matchName: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
        elif (endpoin.find("/") == True):
            # Значение является Namespace/app
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 12*"\xa0" + f"io.kubernetes.pod.namespace: {namespace}" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
        elif (endpoin.find("app:")):
            # Значение является app= в рамках текущего namespace
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
        else:
            # Значение является app= в рамках текущего namespace
            with open("hello.txt", "a") as file:
                file.write(4*"\xa0" + "- toEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
    def ingress (self, namespace, name_policy, number, proto, app, endpoin, port):
        if (endpoin.endswith("/32") == True):
            # Значение является CIDR
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- fromCIDRSet:\n"  + 8*"\xa0" + f"- cidr: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" +  "egress:\n" + 4*"\xa0" + "- {}\n" + "\n---\n")
        elif (endpoin.find("gazprom-neft.local") != -1):
            # Значение является FQDN
            with open(f"{name_policy}.yml", "a") as file:
                file.write("FQDN не может указываться при значении ingress в качестве источника трафика. Необходимо указание по CIDR или по Namespace/app." + f" Ошибка в строчке {number}" + "\n---\n")
        elif (endpoin.find("/") == True):
            # Значение является Namespace/app
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- fromEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 12*"\xa0" + f"io.kubernetes.pod.namespace: {namespace}" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" +  "egress:\n" + 4*"\xa0" + "- {}\n" + "\n---\n")
        elif (endpoin.find("app:")):
            # Значение является app= в рамках текущего namespace
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- fromEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" +  "egress:\n" + 4*"\xa0" + "- {}\n" + "\n---\n")
        else:
            # Значение является app= в рамках текущего namespace
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- fromEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" +  "egress:\n" + 4*"\xa0" + "- {}\n" + "\n---\n")
    def err (self,name_policy, number):
        with open(f"{name_policy}.yml", "a") as file:
            file.write("\n Ошибка чтения значения ingress/egress в строке = " + f"{number}" + 2*"\n")


# In[22]:


class Convert2:
    def yml(self, namespace, name_policy, number, proto, app, endpoin, port, ing_egr, type_endpoin):
        if ing_egr == "egress":
            # Значение является CIDR
            if (type_endpoin == "CIDR"):
                with open(f"{name_policy}.yml", "a") as file:
                    file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toCIDRSet:\n"  + 8*"\xa0" + f"- cidr: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
            # Значение является FQDN
            elif (type_endpoin == "FQDN"):
                with open(f"{name_policy}.yml", "a") as file:
                    file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toFQDNs:\n"  + 8*"\xa0" + f"- matchName: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
            # Значение является Namespace/app
            elif (type_endpoin == "NS"):
                with open(f"{name_policy}.yml", "a") as file:
                    file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin[0]}\n" + 12*"\xa0" + f"io.kubernetes.pod.namespace: {endpoin[1]}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
            # Значение является значением с меткой в рамках текущего namespace
            elif (type_endpoin == "APP"):
                with open(f"{name_policy}.yml", "a") as file:
                    file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"{endpoin[0]}: {endpoin[1]}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
            elif (type_endpoin == "Error"):
                print ("Ошибка в разделении endpoin = " + str(text))
                return "Error"
            else:
                return "Error"


# In[4]:


#Класс для определения значения endpoint
class Valid:
    def valid_endpoint (self, endpoin):
        if (re.search(r'\d+\.+\d+\.+\d+\.+\d+\d+\/32', endpoin) != None):
            return "CIDR"
        elif(re.search(r'\S+\.+gazprom-neft+\.+local', endpoin) != None):
            return "FQDN"
        elif(re.search (r'\D+\/+\D', endpoin) != None):
            return "NS"
        elif(re.search(r'\S+\:+\S', endpoin) != None):
            return "APP"
        else:
            print ("Ошибка в обработке endpoin = " + str(endpoin))
            return "Error"


# In[5]:


class SplitTXT:
    def split_text (self, text, type_endpoin):
        text = text.lower()        
        if (type_endpoin == "namespace"):
            return namespace[text.find('u'):len(text)]
        elif (type_endpoin == "CIDR"):
            return text
        elif (type_endpoin == "FQDN"):
            return text
        elif (type_endpoin == "NS"):
            return re.split(r'/', text)
        elif (type_endpoin == "APP"):
            return re.split(r':', text)
        elif (type_endpoin == "Error"):
            print ("Ошибка в разделении endpoin = " + str(text))
            return "Error"
        else:
            return "Error"


# In[6]:


# Инициализация чтения таблицы
file_path = "cilium.xlsx"
workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
workbook.sheetnames
sheet = workbook.active


# In[7]:


# Ининциализация классов
convert = Convert()
#conv = Convert2()
valid = Valid()
split = SplitTXT()


# In[8]:


# Получение namespace из объединеннной строки под индексом А2x
namespace = sheet["A2"].value
#namespace = namespace.lower()
#len_namespace = len(namespace)
#index_u = namespace.find('u')
#namespace = namespace[index_u:len_namespace]
namespace = split.split_text(namespace, "namespace")


# In[9]:


# Создание номера правила из имени файла
name_policy = file_path[0:file_path.find(".")].lower()


# In[10]:


# Тест регулярки класса valid
#n = 7
#number = sheet["A"+str(n)].value
#proto = sheet["B"+str(n)].value 
#app = sheet["C"+str(n)].value 
#endpoin = sheet["E"+str(n)].value
#port = sheet["F"+str(n)].value
#res_ip = re.search(r'\d+\.+\d+\.+\d+\.+\d+\d+\/32', endpoin)
#res_fqdn = re.search(r'\S+\.+gazprom-neft+\.+local', endpoin)
#res_namesp = re.search (r'\D+\/+\D', endpoin)
#res_app = re.search(r'app:+\S', endpoin)
#print ('В строчке = ' + str(n) + '\nres_fqdn = ' + str(res_fqdn) + '\nres_ip = ' + str(res_ip) + '\nres_namesp = ' + str(res_namesp) + "\nres_app = " + str(res_app) + 2*"\n")
#print ("Endpoin = " + str(endpoin) + "\n" + str(valid.valid_endpoint(endpoin)))


# In[38]:


# Тест класса SplitTXT
n = 4
#number = sheet["A"+str(n)].value
#proto = sheet["B"+str(n)].value 
#app = sheet["C"+str(n)].value 
endpoin = sheet["E"+str(n)].value
#port = sheet["F"+str(n)].value
type_text = str(valid.valid_endpoint(endpoin))
#print (type_text)
endpoin1 = (split.split_text(endpoin , type_text))
print (endpoin1[0])


# In[13]:


# Тест класса Convert2
conv = Convert2()
i = 3
#number = sheet["A"+str(i)].value
#proto = sheet["B"+str(i)].value 
#app = sheet["C"+str(i)].value
#endpoin = sheet["E"+str(i)].value
#port = sheet["F"+str(i)].value
#type_endpoin = str(valid.valid_endpoint(endpoin))
#conv.yml(namespace, name_policy, sheet["A"+str(i)].value, sheet["B"+str(i)].value, sheet["C"+str(i)].value, split.split_text(endpoin , type_text), sheet["F"+str(i)].value, sheet["D"+str(i)].value, type_endpoin)


# In[ ]:





# In[ ]:





# In[ ]:





# In[24]:


# Перебор строк в таблитце до первой пустой строки. Начало с 3 строки таблицы включительно
conv = Convert2()
i = 3
while sheet[f'A{i}'].value is not None:
    number = sheet["A"+str(i)].value
    proto = sheet["B"+str(i)].value 
    app = sheet["C"+str(i)].value
    endpoin = sheet["E"+str(i)].value
    port = sheet["F"+str(i)].value
    ing_egr = sheet["D"+str(i)].value
    type_endpoin = str(valid.valid_endpoint(endpoin))
    #(self, namespace, name_policy, number, proto, app, endpoin, port, ing_egr, type_endpoin)
    conv.yml(namespace, name_policy, number, proto, app, split.split_text(endpoin , type_text), port, ing_egr, type_endpoin)
    i += 1
#    if  sheet["D"+str(i)].value == "ingress":
#        convert.ingress (namespace, name_policy, sheet["A"+str(i)].value, sheet["B"+str(i)].value, sheet["C"+str(i)].value, sheet["E"+str(i)].value, sheet["F"+str(i)].value)
#    elif sheet["D"+str(i)].value == "egress":
#        convert.egress (namespace, name_policy, sheet["A"+str(i)].value, sheet["B"+str(i)].value, sheet["C"+str(i)].value, sheet["E"+str(i)].value, sheet["F"+str(i)].value)
#    else:
#        convert.err (name_policy, sheet["A"+str(i)].value)
#    i += 1



# In[ ]:





# In[ ]:





# In[ ]:





# In[65]:


# Закрытие файла excel таблицы
workbook.close()


# In[ ]:




