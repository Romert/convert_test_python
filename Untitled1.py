#!/usr/bin/env python
# coding: utf-8

# In[ ]:


from openpyxl import load_workbook


# In[61]:


class Convert:
    def egress (self, namespace, name_policy,  number, proto, app, endpoin, port):
        if (endpoin.endswith("/32") == True):
            # Значение является CIDR
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toCIDRSet:\n"  + 8*"\xa0" + f"- cidr: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
        elif (endpoin.find("gazprom-neft.local") == True):
            # Значение является FQDN
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toFQDNs:\n"  + 8*"\xa0" + f"- matchName: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
        elif (endpoin.endswith("/") == True):
            # Значение является Namespace/app
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 12*"\xa0" + f"io.kubernetes.pod.namespace: {namespace}" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
        elif (endpoin.find("app:")):
            # Значение является app= в рамках текущего namespace
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- {}\n" + 2*"\xa0" + "egress:\n" + 4*"\xa0" + "- toEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
        else:
            # Значение является app= в рамках текущего namespace
            with open("hello.txt", "a") as file:
                file.write(4*"\xa0" + "- toEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" + "\n---\n")
    def ingress (self, namespace, name_policy, number, proto, app, endpoin, port):
        if (endpoin.endswith("/32") == True):
            # Значение является CIDR
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- fromCIDRSet:\n"  + 8*"\xa0" + f"- cidr: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" +  "egress:\n" + 4*"\xa0" + "- {}\n" + "\n---\n")
        elif (endpoin.find("gazprom-neft.local") != -1):
            # Значение является FQDN
            with open(f"{name_policy}.yml", "a") as file:
                file.write("FQDN не может указываться при значении ingress в качестве источника трафика. Необходимо указание по CIDR или по Namespace/app." + f" Ошибка в строчке {number}" + "\n---\n")
        elif (endpoin.endswith("/") == True):
            # Значение является Namespace/app
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- fromEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 12*"\xa0" + f"io.kubernetes.pod.namespace: {namespace}" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" +  "egress:\n" + 4*"\xa0" + "- {}\n" + "\n---\n")
        elif (endpoin.find("app:")):
            # Значение является app= в рамках текущего namespace
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- fromEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" +  "egress:\n" + 4*"\xa0" + "- {}\n" + "\n---\n")
        else:
            # Значение является app= в рамках текущего namespace
            with open(f"{name_policy}.yml", "a") as file:
                file.write("apiVersion: cilium.io/v2\n" + "kind: CiliumNetworkPolicy\n" + "metadata:\n" + 2*"\xa0" + f"name: {name_policy}-{number}\n"+ 2*"\xa0" + f"namespace: {namespace}\n" + "spec:\n" + 2*"\xa0" + "endpointSelector:\n" + 4*"\xa0" + "matchLabels:\n" + 6*"\xa0" + f"app: {app}\n" + 2*"\xa0" + "ingress:\n" + 4*"\xa0" + "- fromEndpoints:\n"  + 8*"\xa0" + "- matchLabels:\n" + 12*"\xa0" + f"app: {endpoin}\n" + 6*"\xa0" + "toPorts:\n" + 8*"\xa0" + "- ports:\n" + 12*"\xa0" + f"- port: \"{port}\"\n" + 14*"\xa0" + f"protocol: {proto}\n" +  "egress:\n" + 4*"\xa0" + "- {}\n" + "\n---\n")
    def err (self,name_policy, number):
        with open(f"{name_policy}.yml", "a") as file:
            file.write("\n Ошибка чтения значения ingress/egress в строке = " + f"{number}" + 2*"\n")


# In[62]:


# Инициализация чтения таблицы
file_path = "cilium1.xlsx"
workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
workbook.sheetnames
sheet = workbook.active


# In[63]:


convert = Convert()


# In[64]:


# Получение namespace из объединеннной строки под индексом А2x
namespace = sheet["A2"].value
namespace = namespace.lower()
len_namespace = len(namespace)
index_u = namespace.find('u')
namespace = namespace[index_u:len_namespace]


# In[65]:


# Создание номера правила из имени файла
name_policy = file_path[0:file_path.find(".")].lower()


# In[66]:


# Перебор строк в таблитце до первой пустой строки. Начало с 3 строки таблицы включительно
i = 3
while sheet[f'A{i}'].value is not None:
    if  sheet["D"+str(i)].value == "ingress":
        convert.ingress (namespace, name_policy, sheet["A"+str(i)].value, sheet["B"+str(i)].value, sheet["C"+str(i)].value, sheet["E"+str(i)].value, sheet["F"+str(i)].value)
    elif sheet["D"+str(i)].value == "egress":
        convert.egress (namespace, name_policy, sheet["A"+str(i)].value, sheet["B"+str(i)].value, sheet["C"+str(i)].value, sheet["E"+str(i)].value, sheet["F"+str(i)].value)
    else:
        convert.err (name_policy, sheet["A"+str(i)].value)
    i += 1


# In[24]:


# Закрытие файла excel таблицы
workbook.close()


# In[45]:


test = "FWEFD3421QWE.gazprom-neft.local"
if test.find("gazprom-neft.local") != -1:
    print("work")
else:
    print("error")


# In[ ]:




